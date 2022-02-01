import os
import calendar
from calendar import monthrange
from datetime import timedelta
from time import strftime as time_strftime
from django.shortcuts import render
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.views.generic.list import ListView
from django.contrib.auth.mixins import PermissionRequiredMixin
from openpyxl import Workbook
import openpyxl.drawing.image as openpyxlImage
from tempfile import NamedTemporaryFile
from django.http import HttpResponse
from django.templatetags.static import static
from django.conf import settings

from frequenze.openpyxl_styles import *
from .models import *
from .forms import *
from core.auth_lists import BACKOFFICE
from django.views.decorators.http import require_http_methods
import sys

from .utils import strfdelta_round


class TeacherAgreementCreateView(PermissionRequiredMixin, CreateView):
    permission_required = 'amministrazione.view_teacheragreement'
    model = TeacherAgreement
    form_class = TeacherAgreementModelForm
    template_name = "teacher-agreement/teacher_agreement_create.html"

    def get_success_url(self):
        return reverse('teacher-agreement-list')


def teacher_agreement_detail_view(request, pk):
    permission = request.user.has_perm('amministrazione.view_teacheragreement')
    if permission:
        agreement = get_object_or_404(TeacherAgreement, pk=pk)
        context = {"agreement": agreement}
        return render(request, "teacher-agreement/teacher_agreement_detail.html", context)
    else:
        return render(request, 'missing_auth.html')


class TeacherAgreementUpdateView(PermissionRequiredMixin, UpdateView):
    permission_required = 'amministrazione.view_teacheragreement'
    model = TeacherAgreement
    form_class = TeacherAgreementModelForm
    template_name = "teacher-agreement/teacher_agreement_update.html"

    def get_success_url(self):
        return reverse('teacher-agreement-list')


def display_teacher_agreement(request):
    permission = request.user.has_perm('amministrazione.view_teacheragreement')
    if permission:
        list = TeacherAgreement.objects.all()
        return render(request, 'teacher-agreement/teacher_agreement_list.html', {'list': list})
    else:
        return render(request, 'missing_auth.html')


@require_http_methods(['DELETE'])
def delete_teacher_agreement(request, pk):
    permission = request.user.has_perm('amministrazione.delete_teacheragreement')
    if permission:
        TeacherAgreement.objects.filter(pk=pk).delete()
        agreements = TeacherAgreement.objects.all()
        return render(request, 'teacher-agreement/include/display_teacher_agreement_list.html', {'list': agreements})
    else:
        return render(request, 'missing_auth.html')


########################################################################################################################


def timecard_entry_detail(request, pk, month, year):
    period_start = datetime(year, month, 1)
    if month == 12:
        period_end = datetime(year, month, 31)
    else:
        period_end = datetime(year, month+1, 1)

    employee = Staff.objects.get(pk=pk)
    timecard = TimeCardEntry.objects.filter(employee_id=pk,
                                            entry_date__range=[period_start, period_end]).order_by('entry_date',
                                                                                                   'entry_time')
    previous_month_date_obj = period_start.date() - timedelta(days=1)
    next_month_date_obj = period_start.date() + timedelta(days=31)
    date_nav = {
        'prev_month': previous_month_date_obj.month,
        'next_month': next_month_date_obj.month,
        'prev_year': previous_month_date_obj.year,
        'next_year': next_month_date_obj.year,
    }
    return render(request, 'timecard_entry/timecard_entry_detail.html', context={'pk': pk,
                                                                                 'employee': employee,
                                                                                 "month_name": calendar.month_name[
                                                                                     month],
                                                                                 'timecard': timecard,
                                                                                 'month': month,
                                                                                 'year': year,
                                                                                 'date_nav': date_nav})


def timecard_entry_update(request, *args, **kwargs):
    pk = kwargs['pk']
    if request.method == 'POST':
        entry = get_object_or_404(TimeCardEntry, pk=pk)
        entry.entry_time = request.POST.get('time')
        entry.save(update_fields=['entry_time'])
        return redirect(reverse('timecard-list', kwargs={'pk': entry.employee.id}))
    else:
        entry = get_object_or_404(TimeCardEntry, pk=pk)
    return render(request, 'timecard_entry/timecard_entry_update.html', {'entry': entry})


class TimeCardListView(PermissionRequiredMixin,
                       ListView):
    permission_required = 'amministrazione.view_timecard'
    template_name = 'timecard_entry/timecard_entry_list.html'
    context_object_name = "context"

    def get_queryset(self):
        queryset = {'entries': TimeCardEntry.objects.filter(employee_id=self.kwargs['pk'],
                                                            entry_date__gte=(datetime.today() - timedelta(days=2)))}
        queryset['employee'] = queryset['entries'].last().employee.name + ' ' + queryset['entries'].last().employee.surname
        return queryset

def timecard_list(request, *args, **kwargs):

    entries = TimeCardEntry.objects.filter(employee_id=kwargs['pk'],
                                           entry_date__gte=(datetime.today() - timedelta(days=7))).order_by('-pk')
    context = {'entries': entries}
    if entries:
        context['employee'] = context['entries'].last().employee.name + ' ' + context['entries'].last().employee.surname
    else:
        context = {'employee': get_object_or_404(Staff, pk=kwargs['pk'])}

    context['staff'] = Staff.objects.all()

    if request.method == 'POST':
        pk = request.POST.get('search_pk')
        return redirect(reverse('timecard-list', kwargs={'pk': pk}))
    return render(request, 'timecard_entry/timecard_entry_list.html', context)

def download_timecard(request, pk, month, year):
    reference_date = datetime(day=1, month=month, year=year)
    month_lenght = monthrange(year, month)[1]
    employee = Staff.objects.get(pk=pk)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Timecard"
    characters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    max_width = 25

    # ws1.column_dimensions[characters[char_index]].width = max_width - 5
    header_url = os.path.join(settings.STATIC_ROOT, 'images/header_carta_intestata.png')
    img = openpyxlImage.Image(header_url)
    img.width = 585
    img.height = 51.7
    img.anchor = 'A1'
    ws1.add_image(img)
    ws1['A7'] = f'Timecard di {employee.name.upper()} {employee.surname.upper()}'
    ws1['A7'].font = FONT_DAYS
    ws1['F7'] = f'Mese: ' + datetime.strftime(reference_date, '%B %Y').upper()
    ws1['F7'].font = FONT_DAYS
    ws1['C10'] = ws1['E10'] = 'ENTRATA'
    ws1['C10'].font = ws1['E10'].font = FONT_DAYS
    ws1['C10'].border = ws1['E10'].border = BORDER_HOURS
    ws1['C10'].alignment = ws1['E10'].alignment = ALIGNMENT
    ws1['D10'] = ws1['F10'] = 'USCITA'
    ws1['D10'].font = ws1['F10'].font = FONT_DAYS
    ws1['D10'].border = ws1['F10'].border = BORDER_HOURS
    ws1['D10'].alignment = ws1['F10'].alignment = ALIGNMENT
    ws1['G10'] = 'ORE'
    ws1['G10'].font = FONT_DAYS
    ws1['G10'].border = BORDER_HOURS
    ws1['G10'].alignment = ALIGNMENT
    ws1.column_dimensions['G'].width = 15
    day = 1
    row_index = 11
    column_index = 0
    monthly_attendances = TimeCardEntry.objects.filter(employee_id=pk, entry_date__month=month, entry_date__year=year).order_by('pk')
    time_working_month = timedelta(0, 0)
    while day <= month_lenght:
        ws1[f'{characters[column_index]}{row_index}'] = day
        ws1[f'{characters[column_index]}{row_index}'].font = FONT_HOURS
        ws1[f'{characters[column_index]}{row_index}'].border = BORDER_HOURS
        ws1[f'{characters[column_index]}{row_index}'].alignment = ALIGNMENT
        column_index += 1
        ws1[f'{characters[column_index]}{row_index}'] = datetime.strftime(datetime(day=day, month=month, year=year),
                                                                          "%a")
        i = 1
        while i < 7:
            ws1[f'{characters[i]}{row_index}'].font = FONT_HOURS
            ws1[f'{characters[i]}{row_index}'].border = BORDER_HOURS
            ws1[f'{characters[i]}{row_index}'].alignment = ALIGNMENT
            i += 1
        dayly_attendances = monthly_attendances.filter(entry_date__day=day)
        time_working = 0
        if dayly_attendances:
            column_index += 1
            start_time = 0
            for attendance in dayly_attendances:
                ws1[f'{characters[column_index]}{row_index}'] = datetime.strftime(datetime(day=day,
                                                                                           month=month,
                                                                                           year=year,
                                                                                           hour=attendance.entry_time.hour,
                                                                                           minute=attendance.entry_time.minute),
                                                                                  '%-H:%M')

                if attendance.in_out == 'E':
                    start_time = attendance.entry_time
                elif attendance.in_out == 'U':
                    if start_time:
                        time_working = timedelta(hours=attendance.entry_time.hour - start_time.hour,
                                                 minutes=attendance.entry_time.minute - start_time.minute)
                column_index += 1

        if time_working:
            ws1[f'{characters[6]}{row_index}'] = strfdelta_round(time_working)
            ws1[f'{characters[column_index]}{row_index}'].font = FONT_HOURS
            ws1[f'{characters[column_index]}{row_index}'].border = BORDER_HOURS
            ws1[f'{characters[column_index]}{row_index}'].alignment = ALIGNMENT
            time_working_month += time_working

        column_index = 0
        row_index += 1
        day += 1

    ws1[f'A{row_index + 2}'] = 'Firma: ' + '_' * 25
    ws1[f'A{row_index + 2}'].font = FONT_DAYS
    ws1[f'{characters[5]}{row_index + 2}'] = 'TOTALE'
    ws1[f'{characters[5]}{row_index + 2}'].font = FONT_DAYS
    ws1[f'{characters[5]}{row_index + 2}'].border = BORDER_HOURS
    ws1[f'{characters[5]}{row_index + 2}'].alignment = ALIGNMENT
    ws1[f'{characters[6]}{row_index + 2}'] = strfdelta_round(time_working_month)
    ws1[f'{characters[6]}{row_index + 2}'].font = FONT_DAYS
    ws1[f'{characters[6]}{row_index + 2}'].border = BORDER_HOURS
    ws1[f'{characters[6]}{row_index + 2}'].alignment = ALIGNMENT

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    response = HttpResponse(content=stream, content_type='application/ms-excel', )
    response['Content-Disposition'] = 'attachment; filename=timecard.xlsx'
    return response


def download_all_timecards(request, month, year):
    reference_date = datetime(day=1, month=month, year=year)
    month_lenght = monthrange(year, month)[1]
    staff = Staff.objects.all().order_by('surname')
    wb = Workbook()
    ws = {}
    max_width = 25
    characters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    for employee in staff:
        timecard_found = TimeCardEntry.objects.filter(employee_id=employee.pk,
                                                      entry_date__month=month,
                                                      entry_date__year=year)
        if timecard_found:
            ws[f'{employee.surname} {employee.name}'] = wb.create_sheet()
            ws[f'{employee.surname} {employee.name}'].title = f'{employee.surname} {employee.name}'
            # ws1.column_dimensions[characters[char_index]].width = max_width - 5
            header_url = os.path.join(settings.STATIC_ROOT, 'images/header_carta_intestata.png')
            img = openpyxlImage.Image(header_url)
            img.width = 585
            img.height = 51.7
            img.anchor = 'A1'
            ws[f'{employee.surname} {employee.name}'].add_image(img)
            ws[f'{employee.surname} {employee.name}']['A7'] = f'Timecard di {employee.name.upper()} {employee.surname.upper()}'
            ws[f'{employee.surname} {employee.name}']['A7'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}']['F7'] = f'Mese: ' + datetime.strftime(reference_date, '%B %Y').upper()
            ws[f'{employee.surname} {employee.name}']['F7'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}']['C10'] = ws[f'{employee.surname} {employee.name}']['E10'] = 'ENTRATA'
            ws[f'{employee.surname} {employee.name}']['C10'].font = ws[f'{employee.surname} {employee.name}']['E10'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}']['C10'].border = ws[f'{employee.surname} {employee.name}']['E10'].border = BORDER_HOURS
            ws[f'{employee.surname} {employee.name}']['C10'].alignment = ws[f'{employee.surname} {employee.name}']['E10'].alignment = ALIGNMENT
            ws[f'{employee.surname} {employee.name}']['D10'] = ws[f'{employee.surname} {employee.name}']['F10'] = 'USCITA'
            ws[f'{employee.surname} {employee.name}']['D10'].font = ws[f'{employee.surname} {employee.name}']['F10'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}']['D10'].border = ws[f'{employee.surname} {employee.name}']['F10'].border = BORDER_HOURS
            ws[f'{employee.surname} {employee.name}']['D10'].alignment = ws[f'{employee.surname} {employee.name}']['F10'].alignment = ALIGNMENT
            ws[f'{employee.surname} {employee.name}']['G10'] = 'ORE'
            ws[f'{employee.surname} {employee.name}']['G10'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}']['G10'].border = BORDER_HOURS
            ws[f'{employee.surname} {employee.name}']['G10'].alignment = ALIGNMENT
            ws[f'{employee.surname} {employee.name}'].column_dimensions['G'].width = 15
            day = 1
            row_index = 11
            column_index = 0
            monthly_attendances = TimeCardEntry.objects.filter(employee_id=employee.pk, entry_date__month=month, entry_date__year=year).order_by('pk')
            time_working_month = timedelta(0, 0)
            while day <= month_lenght:
                ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'] = day
                ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'].font = FONT_HOURS
                ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'].border = BORDER_HOURS
                ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'].alignment = ALIGNMENT
                column_index += 1
                ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'] = datetime.strftime(datetime(day=day, month=month, year=year),
                                                                                  "%a")
                i = 1
                while i < 7:
                    ws[f'{employee.surname} {employee.name}'][f'{characters[i]}{row_index}'].font = FONT_HOURS
                    ws[f'{employee.surname} {employee.name}'][f'{characters[i]}{row_index}'].border = BORDER_HOURS
                    ws[f'{employee.surname} {employee.name}'][f'{characters[i]}{row_index}'].alignment = ALIGNMENT
                    i += 1
                dayly_attendances = monthly_attendances.filter(entry_date__day=day)
                time_working = 0
                if dayly_attendances:
                    column_index += 1
                    start_time = 0
                    for attendance in dayly_attendances:
                        ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'] = datetime.strftime(datetime(day=day,
                                                                                                   month=month,
                                                                                                   year=year,
                                                                                                   hour=attendance.entry_time.hour,
                                                                                                   minute=attendance.entry_time.minute),
                                                                                          '%-H:%M')

                        if attendance.in_out == 'E':
                            start_time = attendance.entry_time
                        elif attendance.in_out == 'U':
                            if start_time:
                                time_working = timedelta(hours=attendance.entry_time.hour - start_time.hour,
                                                         minutes=attendance.entry_time.minute - start_time.minute)
                        column_index += 1

                if time_working:
                    ws[f'{employee.surname} {employee.name}'][f'{characters[6]}{row_index}'] = strfdelta_round(time_working)
                    ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'].font = FONT_HOURS
                    ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'].border = BORDER_HOURS
                    ws[f'{employee.surname} {employee.name}'][f'{characters[column_index]}{row_index}'].alignment = ALIGNMENT
                    time_working_month += time_working

                column_index = 0
                row_index += 1
                day += 1

            ws[f'{employee.surname} {employee.name}'][f'A{row_index + 2}'] = 'Firma: ' + '_' * 25
            ws[f'{employee.surname} {employee.name}'][f'A{row_index + 2}'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}'][f'{characters[5]}{row_index + 2}'] = 'TOTALE'
            ws[f'{employee.surname} {employee.name}'][f'{characters[5]}{row_index + 2}'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}'][f'{characters[5]}{row_index + 2}'].border = BORDER_HOURS
            ws[f'{employee.surname} {employee.name}'][f'{characters[5]}{row_index + 2}'].alignment = ALIGNMENT
            ws[f'{employee.surname} {employee.name}'][f'{characters[6]}{row_index + 2}'] = strfdelta_round(time_working_month)
            ws[f'{employee.surname} {employee.name}'][f'{characters[6]}{row_index + 2}'].font = FONT_DAYS
            ws[f'{employee.surname} {employee.name}'][f'{characters[6]}{row_index + 2}'].border = BORDER_HOURS
            ws[f'{employee.surname} {employee.name}'][f'{characters[6]}{row_index + 2}'].alignment = ALIGNMENT

    try:
        remove_me = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(remove_me)
    except:
        pass

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    response = HttpResponse(content=stream, content_type='application/ms-excel', )
    response['Content-Disposition'] = f'attachment; filename=timecard_generale_{month}_{year}.xlsx'
    return response


########################################################################################################################

class WarehouseListView(PermissionRequiredMixin,
                        ListView):
    permission_required = 'amministrazione.view_warehouse'
    queryset = Warehouse.objects.all().order_by("last_movement",
                                                "name")
    template_name = 'warehouse/warehouse_list.html'
    context_object_name = "list"


def warehouse_scan_code(request):
    if request.method == 'POST':
        barcode = request.POST.get('code')
        warehouse = Warehouse.objects.all()
        try:
            object = warehouse.get(barcode=barcode)
            movements = Movement.objects.filter(product=object).order_by('-movement_date',
                                                                         '-id')
            return render(request, 'warehouse/product_detail.html', context={'product': object,
                                                                             'movements': movements})
        except ObjectDoesNotExist:
            return redirect(reverse('warehouse-movement-create', kwargs={'barcode': barcode}))
    else:
        return render(request, 'warehouse/test_htmx.html')


def test_htmx(request):
    return render(request, 'warehouse/test_htmx.html')


def warehouse_movement_create(request, barcode=None):
    if request.method == 'POST':
        post = request.POST
        movement_date = datetime.strptime(post['movement_date'], '%d/%m/%Y')
        product = Warehouse(barcode=barcode,
                            name=post['name'],
                            brand=post['brand'],
                            description=post['description'],
                            department=post['department'],
                            price=post['price'],
                            quantity=post['quantity'],
                            last_movement=movement_date
                            )
        product.save()
        movement = Movement(product=product,
                            movement_type='+',
                            quantity=post['quantity'],
                            movement_date=movement_date)
        movement.save()

        movements = Movement.objects.filter(product=product).order_by('-movement_date')
        return render(request, 'warehouse/product_detail.html', context={'product': product,
                                                                         'movements': movements})

    else:
        movement_form = NewProductMovementModelForm()
        product_form = NewProductModelForm()
        return render(request, 'warehouse/add_product.html', context={'barcode': barcode,
                                                                      'movement_form': movement_form,
                                                                      'product_form': product_form})


def movement_add(request, barcode=None):
    if request.method == 'GET':
        movement_form = MovementModelForm()
        return render(request, 'warehouse/movement_add.html', context={'barcode': barcode,
                                                                       'movement_form': movement_form}
                      )
    else:
        post = request.POST
        product = Warehouse.objects.get(barcode=barcode)
        movement_date = datetime.strptime(post['movement_date'], '%d/%m/%Y')
        movement = Movement(product=product,
                            movement_type=post['movement_type'],
                            quantity=post['quantity'],
                            movement_date=movement_date)
        if post['movement_type'] == '+':
            product.quantity += int(post['quantity'])
        else:
            if product.quantity - int(post['quantity']) < 0:
                movement_form = MovementModelForm(data=post)
                return render(request, 'warehouse/movement_add.html', context={'barcode': barcode,
                                                                               'error': True,
                                                                               'movement_form': movement_form}
                              )
            else:
                product.quantity -= int(post['quantity'])
        movement.save()
        product.save()
        return redirect(reverse('warehouse-list'))
