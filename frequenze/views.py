import datetime

from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.views.generic.list import ListView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from openpyxl.writer.excel import save_virtual_workbook
from core.views import get_secretary_data
from frequenze.forms import *
from anagrafica.models import *
from datetime import timedelta
import calendar
from calendar import monthrange
from django.views.decorators.http import require_http_methods
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import generics
from .serializers import *
from django.contrib import messages

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import textwrap
from tempfile import NamedTemporaryFile
from .openpyxl_styles import *


class CommunicationCreateView(CreateView):
    model = FamilyCommunication
    form_class = CommunicationModelForm
    template_name = "communication/communication_create.html"
    success_url = "/"


class CommunicationUpdateView(UpdateView):
    model = FamilyCommunication
    form_class = CommunicationModelForm
    template_name = "teacher/communication_update.html"

    def get_success_url(self):
        return reverse('course-detail', args=[self.kwargs['course']])


class LessonChoiceView(PermissionRequiredMixin,
                       ListView):
    permission_required = 'frequenze.view_teacherattendance'
    queryset = TeacherAttendance.objects.all().order_by("-attendance_date")[:25]
    paginate_by = 25
    template_name = 'lesson/lesson_index.html'
    context_object_name = "lesson_list"

    def get_context_data(self, **kwargs):
        context = super(LessonChoiceView, self).get_context_data(**kwargs)
        context["course_list"] = Course.objects.all().order_by("name")
        context['training_unit_list'] = TrainingUnit.objects.all().order_by("name")
        return context


class RegisterView(PermissionRequiredMixin,
                   ListView):
    permission_required = 'frequenze.view_teacherattendance'
    queryset = TeacherAttendance.objects.all().order_by("-attendance_date")
    template_name = 'lesson/register.html'
    context_object_name = "lesson_list"

    def get_context_data(self, **kwargs):
        context = super(RegisterView, self).get_context_data(**kwargs)
        reference_date = self.request.GET.get('reference_date', None)
        # course = self.request.GET.get('course', None)
        training_unit = self.request.GET.get('training_unit', None)
        if not reference_date:
            reference_date = datetime.now().date()
        else:
            reference_date = datetime.strptime(reference_date, '%Y-%m-%d')
            reference_date = datetime.date((reference_date))

        register = TeacherAttendance.objects.filter(attendance_date__lte=reference_date,
                                                    training_unit=training_unit).order_by('-attendance_date')
        context['register'] = register
        context['course'] = register[1].training_unit.course.name
        context['reference_date'] = reference_date
        context['training_unit'] = training_unit
        return context


class Dictionary(dict):

    def new(self, key, value):
        self[key] = value

    def sum(self, key, attribute, value):
        self[key][attribute] += value

    def add_attribute(self, key, attribute, value):
        self[key][attribute] = value

    def add_value(self, key, value):
        self[key] = value


########################################################################################################################


def teacher_attendance_create(request, teacher, uf, date, course):
    permission = request.user.has_perm('frequenze.add_teacherattendance')
    if permission:
        check_date = datetime.now()
        while True:
            day_name = datetime.strftime(check_date, '%a')
            if date.lower() == day_name:
                print(check_date.date())
                break
            else:
                check_date = check_date - timedelta(days=1)

        try:
            teacher_obj = Staff.objects.get(pk=teacher)
            training_unit_obj = TrainingUnit.objects.get(pk=uf)
            attendance = TeacherAttendance(teacher=teacher_obj,
                                           training_unit=training_unit_obj,
                                           attendance_date=check_date.date())
            attendance.save()
            # request.session['teacher_attendance_saved'] = f'Ora/e registrata per {teacher_obj} il {check_date.date()}'
            print('Ho salvato!')
        except:
            print('Errore nel salvataggio')
        return redirect('lesson-schedule', course=course)

    else:

        return render(request, 'missing_auth.html')


########################################################################################################################


def student_attendance_course_list_view(request, course, month, year):
    permission = request.user.has_perm('frequenze.view_studentattendance')
    if permission:
        context = Dictionary()
        day = 1
        student_list = Student.objects.filter(course=course, is_withdrawn=False).order_by('number', 'surname', 'name')
        month_lenght = monthrange(year, month)[1]
        days_list = []
        for student in student_list:
            context.new(student, {})
            while day <= month_lenght:
                try:
                    student_attendance = StudentAttendance.objects.get(student=student,
                                                                       attendance_date__day=day,
                                                                       attendance_date__month=month,
                                                                       attendance_date__year=year)
                    context.add_attribute(student, day, student_attendance.event)
                except:
                    context.add_attribute(student, day, "")

                day += 1
            day = 1

        day = 1
        while day <= month_lenght:
            days_list.append(day)
            day += 1

        attendances_list = StudentAttendance.objects.all()

        actual_month_str = f'{year}-{month}-01'
        previous_month_date_obj = datetime.strptime(actual_month_str, '%Y-%m-%d').date() - timedelta(days=1)
        next_month_date_obj = datetime.strptime(actual_month_str, '%Y-%m-%d').date() + timedelta(days=31)
        date_nav = {
            'prev_month': previous_month_date_obj.month,
            'next_month': next_month_date_obj.month,
            'prev_year': previous_month_date_obj.year,
            'next_year': next_month_date_obj.year,
        }
        return render(request, "student/attendance_course_list.html", {"context": context,
                                                                       "month": days_list,
                                                                       "date_nav": date_nav,
                                                                       "month_name": calendar.month_name[month],
                                                                       "attendances": attendances_list,
                                                                       "course": course,
                                                                       "course_name": Course.objects.get(pk=course)})
    else:
        return render(request, 'missing_auth.html')


def student_attendance_create_view(request, course, date):
    permission = request.user.has_perm('frequenze.add_studentattendance')
    if permission:
        add_day = datetime.strptime(date, '%Y-%m-%d').date() + timedelta(days=1)
        sub_day = datetime.strptime(date, '%Y-%m-%d').date() - timedelta(days=1)
        add_week = datetime.strptime(date, '%Y-%m-%d').date() + timedelta(days=7)
        sub_week = datetime.strptime(date, '%Y-%m-%d').date() - timedelta(days=7)
        date_nav = {'add_day': add_day,
                    'sub_day': sub_day,
                    'add_week': add_week,
                    'sub_week': sub_week,
                    }
        if request.method == 'POST':
            formset = StudentAttendanceFormSet(course=course, date=date, data=request.POST)
            if formset.is_valid():
                check_if_already_in = StudentAttendance.objects.filter(student__course=course, attendance_date=date)
                if check_if_already_in:
                    for attendance in check_if_already_in:
                        attendance.delete()
                for form in formset:
                    # student = Student.objects.get(stu)
                    print("Dati: ", form.cleaned_data['student'])
                    attendance = StudentAttendance(student=form.cleaned_data['student'],
                                                   attendance_date=form.cleaned_data['attendance_date'],
                                                   event=form.cleaned_data['event'], )
                    # hours_lost=form.cleaned_data['hours_lost'])
                    attendance.save()

                return redirect(Course.get_absolute_url(course))
            else:
                return render(request, "student/attendance_create.html",
                              {"formset": formset, "course": course, "date_nav": date_nav, "date": date})


        else:

            check_if_already_in = StudentAttendance.objects.filter(student__course=course,
                                                                   attendance_date=date).order_by("student__surname")
            if check_if_already_in:
                formset = StudentAttendanceFormSet(queryset=check_if_already_in, course=course, date=date)
            else:
                formset = StudentAttendanceFormSet(queryset=StudentAttendance.objects.none(), course=course, date=date)

            return render(request, "student/attendance_create.html",
                          {"formset": formset, "course": course, "date_nav": date_nav, "date": date})
    else:
        return render(request, 'missing_auth.html')


class StudentAttendanceObj():
    student_obj = ''
    attendance = ''
    key = 0

    def __init__(self, student):
        self.student_obj = student
        self.attendance = 'P'
        self.key = student.pk


def student_new_attendance_create_view(request, course, date):
    permission = request.user.has_perm('frequenze.add_studentattendance')
    if permission:
        message = ''
        add_day = datetime.strptime(date, '%Y-%m-%d').date() + timedelta(days=1)
        sub_day = datetime.strptime(date, '%Y-%m-%d').date() - timedelta(days=1)
        add_week = datetime.strptime(date, '%Y-%m-%d').date() + timedelta(days=7)
        sub_week = datetime.strptime(date, '%Y-%m-%d').date() - timedelta(days=7)
        date_nav = {'add_day': add_day,
                    'sub_day': sub_day,
                    'add_week': add_week,
                    'sub_week': sub_week,
                    }
        students = Student.objects.filter(course__pk=course, is_withdrawn=False).order_by('number')
        attendance_list = []
        old_events = {}
        update = False
        if StudentAttendance.objects.filter(student__course=course,
                                            attendance_date=date):
            update = True
            for student in students:
                old_attendance = StudentAttendance.objects.get(student__pk=student.pk,
                                                               attendance_date=date)
                attendance = StudentAttendanceObj(student)
                attendance.attendance = old_attendance.event
                old_events[student] = old_attendance.event
                attendance_list.append(attendance)

        else:
            for student in students:
                attendance = StudentAttendanceObj(student)
                attendance_list.append(attendance)

        if request.method == 'POST':
            if update:
                for student in students:
                    attendance = StudentAttendance.objects.get(student=student,
                                                               attendance_date=date)
                    attendance.event = request.POST.get(f'{student.pk}')
                    if old_events[student] != attendance.event:
                        if old_events[student] not in ['R', 'RU']:
                            if attendance.event in ['R', 'RU']:
                                edit_student = Student.objects.get(pk=student.pk)
                                edit_student.delays +=1
                                edit_student.save()
                                # print(f'un ritardo in più per {student}. Ora sono {edit_student.delays}')
                        elif old_events[student] in ['R', 'RU']:
                            if attendance.event not in ['R', 'RU']:
                                edit_student = Student.objects.get(pk=student.pk)
                                edit_student.delays -= 1
                                if edit_student.delays < 0:
                                    edit_student.delays = 0
                                edit_student.save()
                                # print(f'un ritardo in meno per {student}. Ora sono {edit_student.delays}')
                        # if old_events[student] != 'R' and old_events[student] != 'R/U':
                        #     print('un ritardo in più')
                        print(student, old_events[student], attendance.event)
                    attendance.save()
            else:
                for student in students:
                    new_attendance = StudentAttendance(student=student,
                                                       attendance_date=date,
                                                       event=request.POST.get(f'{student.pk}'))
                    new_attendance.save()
                    if new_attendance.event in ['R', 'RU']:
                        edit_student = Student.objects.get(pk=student.pk)
                        edit_student.delays += 1
                        edit_student.save()
            attendance_list.clear()
            for student in students:
                old_attendance = StudentAttendance.objects.get(student__pk=student.pk,
                                                               attendance_date=date)
                attendance = StudentAttendanceObj(student)
                attendance.attendance = old_attendance.event
                attendance_list.append(attendance)
            message = 'Presenze salvate!'
            return redirect('course-detail', course)
        date_obj = datetime.strptime(date, "%Y-%m-%d").date()
        return render(request, "student/new_attendance_create.html",
                      {"attendance_list": attendance_list,
                       "course": course,
                       "date_nav": date_nav,
                       "date_obj": date_obj,
                       "date": date,
                       "message": message})
    else:
        return render(request, 'missing_auth.html')


def reset_delays(request):
    permission = request.user.has_perm('frequenze.add_studentattendance')
    if permission:
        students = Student.objects.all()
        for student in students:
            student.delays = 0
            student.save()
        return redirect('index')
    else:
        return render(request, 'missing_auth.html')


def count_delays(request):
    permission = request.user.has_perm('frequenze.add_studentattendance')
    if permission:
        ref_date = datetime(2021, 12, 10)
        attendances = StudentAttendance.objects.filter(attendance_date__gte=ref_date).order_by('student', 'attendance_date')
        students = Student.objects.all()
        for attendance in attendances:
            if attendance.event in ['R', 'RU']:
                edit_delays = students.get(pk=attendance.student.pk)
                edit_delays.delays += 1
                edit_delays.save()
        return redirect('index')
    else:
        return render(request, 'missing_auth.html')
########################################################################################################################


class NotesListView(PermissionRequiredMixin,
                    ListView):
    permission_required = 'frequenze.view_note'
    queryset = Note.objects.all().order_by("-created_at")
    paginate_by = 25
    template_name = 'note/note_list.html'
    context_object_name = "list"


class NotesCreateView(LoginRequiredMixin, CreateView):
    permission_required = 'frequenze.add_note'
    model = Note
    form_class = NoteModelForm
    template_name = "note/note_create.html"
    success_url = "/"

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.created_by = self.request.user
        self.object.save()
        return super().form_valid(form)


class NoteUpdateView(LoginRequiredMixin, UpdateView):
    model = Note
    form_class = NoteModelForm
    template_name = "note/note_update.html"

    def get_success_url(self):
        return reverse('note-list')


def notes_detail_view(request, pk):
    permission = request.user.has_perm('frequenze.view_note')
    if permission:
        note = get_object_or_404(Note, pk=pk)
        if request.method == 'POST':

            note.reminder = not request.POST.get('reminder', False)
            note.solved = request.POST.get('solved', False)
            note.save()
            return redirect('index')
        else:
            context = {"note": note}
            return render(request, "note/note_detail.html", context)
    else:
        return render(request, 'missing_auth.html')


class NotesDeleteView(LoginRequiredMixin, DeleteView):
    permission_required = 'frequenze.delete_note'
    model = Note
    template_name = "note/note_confirm_delete.html"
    success_url = "note/note_list.html"


########################################################################################################################


class CommunicationListView(PermissionRequiredMixin,
                            ListView):
    permission_required = 'anagrafica.view_course'
    paginate_by = 25
    template_name = 'communication/communication_list.html'
    context_object_name = "list"

    def get_queryset(self):
        queryset = FamilyCommunication.objects.filter(course=self.kwargs['course']).order_by('-year', '-month',
                                                                                             'course')
        return queryset

    def get_context_data(self, *, object_list=None, **kwargs):
        course = self.kwargs['course']
        queryset = FamilyCommunication.objects.filter(course=course).order_by('-year', '-month', 'course')

        return {'course': course, 'list': queryset}


def communication_detail_view(request, pk):
    permission = request.user.has_perm('anagrafica.view_course')
    if permission:
        communication = get_object_or_404(FamilyCommunication, pk=pk)
        context = {"communication": communication, 'course': communication.course.pk}
        return render(request, "communication/communication_detail.html", context)
    else:
        return render(request, 'missing_auth.html')


########################################################################################################################


class LessonCreateView(PermissionRequiredMixin, CreateView):
    permission_required = 'frequenze.add_lesson'
    model = Lesson
    form_class = LessonModelForm
    template_name = "lesson/lesson_create.html"
    success_url = "/"


class LessonAddView(PermissionRequiredMixin, CreateView):
    permission_required = 'frequenze.add_lesson'
    model = Lesson
    form_class = LessonModelForm
    template_name = "lesson/lesson_add.html"
    success_url = "/"


def lesson_add(request, course):
    permission = request.user.has_perm('frequenze.add_lesson')

    if permission:
        if request.method == 'POST':
            form = LessonModelForm(data=request.POST, course=course)
            if form.is_valid():
                lesson = Lesson(day=form.cleaned_data['day'],
                                hour=form.cleaned_data['hour'],
                                training_unit=form.cleaned_data['training_unit'],
                                classroom=form.cleaned_data['classroom'])
                try:
                    lesson.save()
                except:
                    print('errore nel salvataggio!')
                return redirect('lesson-schedule', course=course)
            else:

                error_message = ''
                for error in form.non_field_errors():
                    messages.error(request, error)

                return redirect('lesson-schedule', course=course)
                # HttpResponseRedirect(reverse('lesson-schedule', kwargs={'course':course}))
                # redirect('lesson-schedule', course=course)
                # return render(request, "lesson/lesson_add.html",
                #              {"form": form, "course": course})
        else:
            form = LessonModelForm(course=course)
            return render(request, "lesson/lesson_add.html", {"form": form, "course": course})
    else:
        return render(request, 'missing_auth.html')


def new_lesson_add(request, course, ref_date):
    permission = request.user.has_perm('frequenze.add_lesson')
    ref_date = datetime.strptime(ref_date, '%Y-%m-%d').date()
    week_days = {
        'Lun': 'lunedì',
        'Mar': 'martedì',
        'Mer': 'mercoledì',
        'Gio': 'giovedì',
        'Ven': 'venerdì'}
    if permission:
        if request.method == 'POST':
            form = LessonWithDateModelForm(data=request.POST, course=course, ref_date=ref_date)
            if form.is_valid():
                selected_day = week_days[form.cleaned_data['day']]
                while True:
                    if ref_date.strftime("%A") == selected_day:
                        selected_day = ref_date
                        break
                    else:
                        ref_date = ref_date + timedelta(days=1)
                lesson = LessonWithDate(lesson_date=selected_day,
                                        hour=form.cleaned_data['hour'],
                                        training_unit=form.cleaned_data['training_unit'],
                                        classroom=form.cleaned_data['classroom'])
                try:
                    lesson.save()
                except:
                    print('errore nel salvataggio!')
                return redirect('new-lesson-schedule', course=course, ref_date=ref_date)
            else:
                error_message = ''
                for error in form.non_field_errors():
                    messages.error(request, error)
                return redirect('new-lesson-schedule', course=course, ref_date=ref_date)
        else:
            form = LessonWithDateModelForm(course=course, ref_date=ref_date)
            return render(request, "lesson_with_date/lesson_with_date_add.html", {"form": form, "course": course, 'ref_date': ref_date})
    else:
        return render(request, 'missing_auth.html')


def lesson_choose_course(request):
    course_list = Course.objects.filter(is_finished=False)
    # if 'schedule_form_error' in request.session:
    #     del request.session['schedule_form_error']
    # elif 'teacher_attendance_saved' in request.session:
    #     del request.session['teacher_attendance_saved']
    return render(request, 'lesson/lesson_list.html', {'course_list': course_list})


def new_lesson_choose_course(request):
    course_list = Course.objects.filter(is_finished=False)
    ref_date = datetime.now()
    return render(request, 'lesson_with_date/lesson_with_date_list.html', {'course_list': course_list,
                                                                           'ref_date': ref_date.date(),
                                                                           'next_week': (ref_date+timedelta(weeks=1)).date()})


def lesson_schedule(request, course):
    complete_schedule = Lesson.objects.filter(training_unit__course=course)

    schedule = {'hours': ['08.00-9.00',
                          '09.00-10.00',
                          '10.00-11.00',
                          '11.00-12.00',
                          '12.00-13.00',
                          '13.00-14.00',
                          '14.00-15.00',
                          '15.00-16.00',
                          '16.00-17.00'],
                'days': ['Lun',
                         'Mar',
                         'Mer',
                         'Gio',
                         'Ven'
                         ]
                }

    count = 1
    for day in schedule['days']:
        lessons_queryset = complete_schedule.filter(day__exact=day).order_by('hour')
        schedule[day] = {}
        while count < 10:
            try:
                schedule[day][count] = lessons_queryset.get(hour=count)
            except:
                schedule[day][count] = ''

            count += 1
        count = 1
    try:
        course_name = Course.objects.get(pk=course).name
    except:
        course_name = ''

    if request.method == 'POST':
        # del request.session['teacher_attendance_saved']
        form = LessonModelForm(data=request.POST, course=course)
        if form.is_valid():
            lesson = Lesson(day=form.cleaned_data['day'],
                            hour=form.cleaned_data['hour'],
                            training_unit=form.cleaned_data['training_unit'])
            lesson.save()
            # if 'schedule_form_error' in request.session:
            #     del request.session['schedule_form_error']
            return redirect('lesson-schedule', course=course)
        else:
            # request.session['schedule_form_error'] = form.errors
            return redirect('lesson-schedule', course=course)

    else:
        form = LessonModelForm(course=course)
        return render(request, 'lesson/lesson_schedule.html', {"form": form,
                                                               'schedule': schedule,
                                                               'course': course,
                                                               'course_name': course_name})


def new_lesson_schedule(request, course, ref_date):
    ref_date = datetime.strptime(ref_date, '%Y-%m-%d').date()
    while True:
        if ref_date.strftime("%A") == 'lunedì':
            week_begins = ref_date
            week_ends = ref_date + timedelta(days=4)
            break
        else:
            ref_date = ref_date - timedelta(days=1)
    # print(week_begins, week_ends)
    complete_schedule = LessonWithDate.objects.filter(training_unit__course=course,
                                                      lesson_date__range=[week_begins, week_ends]).order_by('lesson_date', 'hour')
    schedule = {'hours': ['08.00-9.00',
                          '09.00-10.00',
                          '10.00-11.00',
                          '11.00-12.00',
                          '12.00-13.00',
                          '13.00-14.00',
                          '14.00-15.00',
                          '15.00-16.00',
                          '16.00-17.00'],
                'days': ['Lun',
                         'Mar',
                         'Mer',
                         'Gio',
                         'Ven'
                         ]
                }
    day_numbers = {}
    for day in schedule['days']:
        day_numbers[day] = f'{ref_date.day} {ref_date.strftime("%B")}'
        ref_date = ref_date + timedelta(days=1)
    ref_date = ref_date - timedelta(days=5)
    selected_days = {
        'lunedì': 'Lun',
        'martedì': 'Mar',
        'mercoledì': 'Mer',
        'giovedì': 'Gio',
        'venerdì': 'Ven'}

    count = 1
    for day in schedule['days']:
        schedule[day] = {}
        while count < 10:
            schedule[day][count] = ''

            count += 1
        count = 1

    for lesson in complete_schedule:
        selected_day = selected_days[lesson.lesson_date.strftime("%A")]
        if selected_day not in schedule.keys():
            schedule[selected_day] = {}
        schedule[selected_day][lesson.hour] = lesson

    # print(schedule)
    try:
        course_name = Course.objects.get(pk=course).name
    except:
        course_name = ''

    if request.method == 'POST':
        # del request.session['teacher_attendance_saved']
        form = LessonWithDateModelForm(data=request.POST, course=course, ref_date=ref_date)
        if form.is_valid():
            # lesson = LessonWithDate(day=form.cleaned_data['day'],
            #                 hour=form.cleaned_data['hour'],
            #                 training_unit=form.cleaned_data['training_unit'])
            # lesson.save()
            # if 'schedule_form_error' in request.session:
            #     del request.session['schedule_form_error']
            return redirect('new-lesson-schedule', course=course)
        else:
            # request.session['schedule_form_error'] = form.errors
            return redirect('new-lesson-schedule', course=course)

    else:
        form = LessonWithDateModelForm(course=course, ref_date=ref_date)
        return render(request, 'lesson_with_date/lesson_with_date_schedule.html',
                      {"form": form,
                       'schedule': schedule,
                       'course': course,
                       'ref_date': week_begins,
                       'week_forward': week_begins + timedelta(weeks=1),
                       'week_backward': week_begins - timedelta(weeks=1),
                       'course_name': course_name,
                       'day_numbers': day_numbers})


def send_schedule(request, course):
    lessons_to_send = Lesson.objects.filter(training_unit__course=course).order_by('day', 'hour')
    lessons_already_sent = LessonSend.objects.filter(training_unit__course=course)

    try:
        for lesson in lessons_already_sent:
            lesson.delete()
    except:
        print('Nessun orario già inviato')

    for lesson in lessons_to_send:
        send_lesson = LessonSend(day=lesson.get_day_display(),
                                 hour=lesson.get_hour_display(),
                                 training_unit=lesson.training_unit)
        send_lesson.save()

    course_obj = Course.objects.get(pk=course)

    class LessonObj:
        def __init__(self, pk, day, hour, training_unit):
            self.pk = pk
            self.day = day
            self.hour = hour
            self.training_unit = training_unit

    list_obj = []

    for day in ['Lun', 'Mar', 'Mer', 'Gio', 'Ven']:
        lessons = lessons_to_send.filter(day=day)

        for l in lessons:
            obj = LessonObj(l.pk, l.get_day_display(), l.get_hour_display(),
                            f'UF{l.training_unit.code} - {l.training_unit.name}')
            list_obj.append(obj)

    return render(request, 'lesson/send_schedule.html', {'schedule': list_obj, 'course': course_obj})


def send_schedule_all_courses(request):
    lessons_to_send = Lesson.objects.all().order_by('training_unit__course', 'day', 'hour')
    lessons_already_sent = LessonSend.objects.all()

    try:
        for lesson in lessons_already_sent:
            lesson.delete()
    except:
        print('Nessun orario già inviato')

    for lesson in lessons_to_send:
        send_lesson = LessonSend(day=lesson.get_day_display(),
                                 hour=lesson.get_hour_display(),
                                 training_unit=lesson.training_unit)
        send_lesson.save()

    course_list = Course.objects.filter(is_finished=False)
    return render(request, 'lesson/lesson_list.html', {'course_list': course_list,
                                                       'message': 'Il nuovo orario è stato inviato a tutti i corsi'})


def new_send_schedule_all_courses(request, ref_date):
    lessons_already_sent = LessonSend.objects.all()
    ref_date = datetime.strptime(ref_date, '%Y-%m-%d').date()
    while True:
        if ref_date.strftime("%A") == 'lunedì':
            break
        else:
            ref_date = ref_date - timedelta(days=1)
    date_check = datetime.now().date()
    while True:
        if date_check.strftime("%A") == 'lunedì':
            break
        else:
            date_check = date_check - timedelta(days=1)
    if date_check == ref_date:
        message = 'Hai inviato a tutti l`orario di questa settimana'
    else:
        message = 'Hai inviato a tutti l`orario della prossima settimana'
    date_range = [ref_date, ref_date + timedelta(days=4)]
    lessons_to_send = LessonWithDate.objects.filter(lesson_date__range=date_range).order_by('training_unit__course', 'lesson_date', 'hour')
    days = ['lunedì',
            'martedì',
            'mercoledì',
            'giovedì',
            'venerdì'
            ]
    try:
        lessons_already_sent.delete()
    except:
        print('Nessun orario già inviato')

    for lesson in lessons_to_send:
        send_lesson = LessonSend(day=lesson.lesson_date.strftime("%A").capitalize(),
                                 hour=lesson.get_hour_display(),
                                 training_unit=lesson.training_unit)
        send_lesson.save()

    course_list = Course.objects.filter(is_finished=False)
    return render(request, 'lesson_with_date/lesson_with_date_list.html', {'course_list': course_list,
                                                                           'ref_date': date_check,
                                                                           'next_week': (ref_date + timedelta(
                                                                               weeks=1)),
                                                                           'message': message})


@require_http_methods(['DELETE'])
def delete_lesson(request, pk):
    permission = request.user.has_perm('frequenze.delete_lesson')
    if permission:
        # if 'schedule_form_error' in request.session:
        #     del request.session['schedule_form_error']
        course = Lesson.objects.get(pk=pk).training_unit.course.pk
        Lesson.objects.get(pk=pk).delete()
        return redirect('lesson-schedule', course=course)
    else:
        return render(request, 'missing_auth.html')


@require_http_methods(['DELETE'])
def new_delete_lesson(request, pk, ref_date):
    permission = request.user.has_perm('frequenze.delete_lesson')
    if permission:
        # if 'schedule_form_error' in request.session:
        #     del request.session['schedule_form_error']
        course = LessonWithDate.objects.get(pk=pk).training_unit.course.pk
        LessonWithDate.objects.get(pk=pk).delete()
        return redirect('new-lesson-schedule', course=course, ref_date=ref_date)
    else:
        return render(request, 'missing_auth.html')


@require_http_methods(['DELETE'])
def new_delete_lesson_to_index(request, pk):
    permission = request.user.has_perm('frequenze.delete_lesson')
    if permission:
        lesson = LessonWithDate.objects.get(pk=pk)
        message = f'{lesson.training_unit} delle {lesson.get_hour_display()} cancellata'
        lesson.delete()
        action = 'delete'
        context = {'user_data': {'secretary_data': get_secretary_data(message, action)}}
        return render(request, './index_include/secretary_lessons_data.html', context)
    else:
        return render(request, 'missing_auth.html')


def new_validate_lesson_to_index(request, pk):
    permission = request.user.has_perm('frequenze.delete_lesson')
    if permission:
        lesson = LessonWithDate.objects.get(pk=pk)
        teacher = lesson.training_unit.contract_related_unit.get().employee
        validate_lesson = LessonDone(lesson_date=lesson.lesson_date,
                                     hour=lesson.hour,
                                     training_unit=lesson.training_unit,
                                     teacher=teacher,
                                     related_lesson=lesson
                                     )
        validate_lesson.save()
        message = f'{validate_lesson.training_unit} delle {validate_lesson.get_hour_display()} validata'
        action = 'validate'
        context = {'user_data': {'secretary_data': get_secretary_data(message, action)}}
        return render(request, './index_include/secretary_lessons_data.html', context)
    else:
        return render(request, 'missing_auth.html')


class ScheduleCourseAPIView(generics.ListAPIView):
    """
    API endpoint
    """

    serializer_class = ScheduleSerializer

    permission_classes = [IsAuthenticated]

    def get_queryset(self):

        class LessonObj:
            def __init__(self, pk, day, hour, training_unit):
                self.pk = pk
                self.day = day
                self.hour = hour
                self.training_unit = training_unit

        course = self.kwargs['course']
        queryset = LessonSend.objects.filter(training_unit__course=course)
        list_obj = []

        for day in ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì']:
            lessons = queryset.filter(day=day)
            for l in lessons:
                obj = LessonObj(l.pk, l.day, l.hour, l.training_unit)
                list_obj.append(obj)
        return list_obj


def general_schedule(request):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "orario settimanale"
    worksheets_dict = {}
    lessons = Lesson.objects.all().order_by('training_unit__course')
    courses_list = ['']
    schedule = {}
    hours = ['08.00-09.00',
             '09.00-10.00',
             '10.00-11.00',
             '11.00-12.00',
             '12.00-13.00',
             '13.00-14.00',
             '14.00-15.00',
             '15.00-16.00',
             '16.00-17.00']
    hours_id_list = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    days = ['Lun',
            'Mar',
            'Mer',
            'Gio',
            'Ven'
            ]

    for lesson in lessons:
        if lesson.training_unit.course.name not in courses_list:
            courses_list.append(lesson.training_unit.course.name)
    characters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    max_width = 25
    char_index = 0
    for course in courses_list:
        i = 1
        day_index = 0
        hour_index = 0
        if char_index == 0:
            max_width = 15
            ws1[f'{characters[char_index]}{i}'] = ''
            i += 1
            insert_day = True
            while True:
                if insert_day:
                    ws1[f'{characters[char_index]}{i}'] = days[day_index]
                    ws1[f'{characters[char_index]}{i}'].font = FONT_DAYS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_DAYS
                    ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT
                    ws1[f'{characters[char_index]}{i}'].fill = FILL_DAYS
                    insert_day = False
                else:
                    ws1[f'{characters[char_index]}{i}'] = hours[hour_index]
                    ws1[f'{characters[char_index]}{i}'].font = FONT_HOURS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_HOURS
                    ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT
                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                i += 1

        else:
            ws1[f'{characters[char_index]}{i}'] = course
            ws1[f'{characters[char_index]}{i}'].font = FONT_COURSES
            ws1[f'{characters[char_index]}{i}'].fill = FILL_COURSES
            ws1[f'{characters[char_index]}{i}'].border = BORDER_COURSES
            ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT
            max_width = 25
            i += 1
            insert_day = True
            while True:

                if insert_day:
                    ws1[f'{characters[char_index]}{i}'] = ''
                    ws1[f'{characters[char_index]}{i}'].fill = FILL_DAYS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_DAYS
                    insert_day = False
                else:
                    try:
                        lesson = lessons.get(training_unit__course__name=course,
                                             day=days[day_index],
                                             hour=hours_id_list[hour_index])

                        if lesson.training_unit.kind == 'L':
                            ws1[f'{characters[char_index]}{i}'].fill = FILL_L
                        elif lesson.training_unit.kind == 'U':
                            ws1[f'{characters[char_index]}{i}'].fill = FILL_U
                        elif lesson.training_unit.kind == 'T':
                            ws1[f'{characters[char_index]}{i}'].fill = FILL_T
                        uf = textwrap.shorten(lesson.training_unit.name,
                                              width=15,
                                              break_long_words=False,
                                              placeholder='')
                        while uf[-1] == ' ':
                            uf = uf[:-1]
                        if uf[-1] == ',':
                            uf = uf[:-1]
                        classroom = f'{lesson.classroom.schedule_name}'
                        if len(classroom) == 1:
                            classroom = f'Aula {classroom}'
                        teacher = lesson.get_teacher()
                        lesson = f'{uf} | {teacher} | {classroom}'
                        ws1[f'{characters[char_index]}{i}'] = lesson
                        if max_width < len(lesson):
                            max_width = len(lesson)
                    except:
                        ws1[f'{characters[char_index]}{i}'] = ''

                    ws1[f'{characters[char_index]}{i}'].font = FONT_LESSONS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_LESSONS
                    ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT

                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                i += 1
        ws1.column_dimensions[characters[char_index]].width = max_width - 5
        if char_index > 0:
            worksheets_dict[char_index + 1] = wb.create_sheet()
            worksheets_dict[char_index + 1].title = course
            max_width = 15
            worksheets_dict[char_index + 1][f'{characters[char_index]}{i}'] = ''
            single_sheet_i = 1
            single_sheet_char_index = 0
            day_index = 0
            hour_index = 0
            worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = ''
            single_sheet_i += 1
            insert_day = True
            while True:
                if insert_day:
                    worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = days[
                        day_index]
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_DAYS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_DAYS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_DAYS
                    insert_day = False
                else:
                    worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = hours[
                        hour_index]
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_HOURS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_HOURS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT
                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                single_sheet_i += 1
            single_sheet_char_index += 1
            single_sheet_i = 1
            day_index = 0
            hour_index = 0
            worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = course
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_COURSES
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_COURSES
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_COURSES
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT
            max_width = 25
            single_sheet_i += 1
            insert_day = True
            while True:

                if insert_day:
                    worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = ''
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_DAYS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_DAYS
                    insert_day = False
                else:
                    try:
                        lesson = lessons.get(training_unit__course__name=course,
                                             day=days[day_index],
                                             hour=hours_id_list[hour_index])

                        if lesson.training_unit.kind == 'L':
                            worksheets_dict[char_index + 1][
                                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_L
                        elif lesson.training_unit.kind == 'U':
                            worksheets_dict[char_index + 1][
                                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_U
                        elif lesson.training_unit.kind == 'T':
                            worksheets_dict[char_index + 1][
                                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_T
                        uf = textwrap.shorten(lesson.training_unit.name,
                                              width=15,
                                              break_long_words=True,
                                              placeholder='')
                        while uf[-1] == ' ':
                            uf = uf[:-1]
                        if uf[-1] == ',':
                            uf = uf[:-1]
                        classroom = f'{lesson.classroom.schedule_name}'
                        if len(classroom) == 1:
                            classroom = f'Aula {classroom}'
                        teacher = lesson.get_teacher()
                        lesson = f'{uf} | {teacher} | {classroom}'
                        worksheets_dict[char_index + 1][
                            f'{characters[single_sheet_char_index]}{single_sheet_i}'] = lesson
                        if max_width < len(lesson):
                            max_width = len(lesson)
                    except:
                        worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = ''

                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_LESSONS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_LESSONS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT

                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                single_sheet_i += 1
            worksheets_dict[char_index + 1].column_dimensions[characters[single_sheet_char_index]].width = max_width - 5
            ws1 = wb.active
        char_index += 1
    ws1.row_dimensions[1].height = 25
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    response = HttpResponse(content=stream, content_type='application/ms-excel', )
    response['Content-Disposition'] = 'attachment; filename=orario_settimanale.xlsx'
    return response



def new_general_schedule(request, ref_date):
    ref_date = datetime.strptime(ref_date, '%Y-%m-%d').date()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "orario settimanale"
    worksheets_dict = {}
    lessons = LessonWithDate.objects.all().order_by('training_unit__course')
    courses_list = ['']
    schedule = {}
    hours = ['08.00-09.00',
             '09.00-10.00',
             '10.00-11.00',
             '11.00-12.00',
             '12.00-13.00',
             '13.00-14.00',
             '14.00-15.00',
             '15.00-16.00',
             '16.00-17.00']
    hours_id_list = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    while True:
        if ref_date.strftime("%A") == 'lunedì':
            break
        else:
            ref_date = ref_date - timedelta(days=1)
    # date_range = [ref_date, ref_date+timedelta(days=4)]
    # useful_lessons = LessonWithDate.objects.filter(lesson_date__range=date_range)
    days = ['lunedì',
            'martedì',
            'mercoledì',
            'giovedì',
            'venerdì'
            ]
    days_n = []
    for i in range(len(days)):
        days_n.append(str((ref_date + timedelta(days=i)).day))

    for lesson in lessons:

        if lesson.training_unit.course.name not in courses_list:
            courses_list.append(lesson.training_unit.course.name)
    characters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    max_width = 25
    char_index = 0
    for course in courses_list:
        i = 1
        day_index = 0
        hour_index = 0
        if char_index == 0:
            max_width = 15
            ws1[f'{characters[char_index]}{i}'] = ''
            i += 1
            insert_day = True
            while True:
                if insert_day:
                    ws1[f'{characters[char_index]}{i}'] = f'{days[day_index][:3].upper()} {days_n[day_index]}'
                    ws1[f'{characters[char_index]}{i}'].font = FONT_DAYS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_DAYS
                    ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT
                    ws1[f'{characters[char_index]}{i}'].fill = FILL_DAYS
                    insert_day = False
                else:
                    ws1[f'{characters[char_index]}{i}'] = hours[hour_index]
                    ws1[f'{characters[char_index]}{i}'].font = FONT_HOURS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_HOURS
                    ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT
                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                i += 1

        else:
            ws1[f'{characters[char_index]}{i}'] = course
            ws1[f'{characters[char_index]}{i}'].font = FONT_COURSES
            ws1[f'{characters[char_index]}{i}'].fill = FILL_COURSES
            ws1[f'{characters[char_index]}{i}'].border = BORDER_COURSES
            ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT
            max_width = 25
            i += 1
            insert_day = True
            while True:

                if insert_day:
                    ws1[f'{characters[char_index]}{i}'] = ''
                    ws1[f'{characters[char_index]}{i}'].fill = FILL_DAYS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_DAYS
                    insert_day = False
                else:
                    try:
                        selected_day = ''
                        counter = 0
                        while True:
                            if ref_date.strftime("%A") == days[day_index]:
                                selected_day = ref_date
                                ref_date = ref_date - timedelta(days=counter)
                                break
                            else:
                                counter+=1
                                ref_date = ref_date + timedelta(days=1)
                        lesson = lessons.get(training_unit__course__name=course,
                                             lesson_date=selected_day,
                                             hour=hours_id_list[hour_index])

                        if lesson.training_unit.kind == 'L':
                            ws1[f'{characters[char_index]}{i}'].fill = FILL_L
                        elif lesson.training_unit.kind == 'U':
                            ws1[f'{characters[char_index]}{i}'].fill = FILL_U
                        elif lesson.training_unit.kind == 'T':
                            ws1[f'{characters[char_index]}{i}'].fill = FILL_T
                        uf = textwrap.shorten(lesson.training_unit.name,
                                              width=15,
                                              break_long_words=False,
                                              placeholder='')
                        while uf[-1] == ' ':
                            uf = uf[:-1]
                        if uf[-1] == ',':
                            uf = uf[:-1]
                        classroom = f'{lesson.classroom.schedule_name}'
                        if len(classroom) == 1:
                            classroom = f'Aula {classroom}'
                        teacher = lesson.get_teacher()
                        lesson = f'{uf} | {teacher} | {classroom}'
                        ws1[f'{characters[char_index]}{i}'] = lesson
                        if max_width < len(lesson):
                            max_width = len(lesson)
                    except:
                        ws1[f'{characters[char_index]}{i}'] = ''

                    ws1[f'{characters[char_index]}{i}'].font = FONT_LESSONS
                    ws1[f'{characters[char_index]}{i}'].border = BORDER_LESSONS
                    ws1[f'{characters[char_index]}{i}'].alignment = ALIGNMENT

                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                i += 1
        ws1.column_dimensions[characters[char_index]].width = max_width - 5
        if char_index > 0:
            worksheets_dict[char_index + 1] = wb.create_sheet()
            worksheets_dict[char_index + 1].title = course
            max_width = 15
            worksheets_dict[char_index + 1][f'{characters[char_index]}{i}'] = ''
            single_sheet_i = 1
            single_sheet_char_index = 0
            day_index = 0
            hour_index = 0
            worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = ''
            single_sheet_i += 1
            insert_day = True
            while True:
                if insert_day:
                    worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = f'{days[day_index][:3].upper()} {days_n[day_index]}'
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_DAYS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_DAYS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_DAYS
                    insert_day = False
                else:
                    worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = hours[
                        hour_index]
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_HOURS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_HOURS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT
                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                single_sheet_i += 1
            single_sheet_char_index += 1
            single_sheet_i = 1
            day_index = 0
            hour_index = 0
            worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = course
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_COURSES
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_COURSES
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_COURSES
            worksheets_dict[char_index + 1][
                f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT
            max_width = 25
            single_sheet_i += 1
            insert_day = True
            while True:

                if insert_day:
                    worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = ''
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_DAYS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_DAYS
                    insert_day = False
                else:
                    try:
                        selected_day = ''
                        counter = 0
                        while True:
                            if ref_date.strftime("%A") == days[day_index]:
                                selected_day = ref_date
                                ref_date = ref_date - timedelta(days=counter)
                                break
                            else:
                                counter += 1
                                ref_date = ref_date + timedelta(days=1)
                        lesson = lessons.get(training_unit__course__name=course,
                                             lesson_date=selected_day,
                                             hour=hours_id_list[hour_index])

                        if lesson.training_unit.kind == 'L':
                            worksheets_dict[char_index + 1][
                                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_L
                        elif lesson.training_unit.kind == 'U':
                            worksheets_dict[char_index + 1][
                                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_U
                        elif lesson.training_unit.kind == 'T':
                            worksheets_dict[char_index + 1][
                                f'{characters[single_sheet_char_index]}{single_sheet_i}'].fill = FILL_T
                        uf = textwrap.shorten(lesson.training_unit.name,
                                              width=15,
                                              break_long_words=True,
                                              placeholder='')
                        while uf[-1] == ' ':
                            uf = uf[:-1]
                        if uf[-1] == ',':
                            uf = uf[:-1]
                        classroom = f'{lesson.classroom.schedule_name}'
                        if len(classroom) == 1:
                            classroom = f'Aula {classroom}'
                        teacher = lesson.get_teacher()
                        lesson = f'{uf} | {teacher} | {classroom}'
                        worksheets_dict[char_index + 1][
                            f'{characters[single_sheet_char_index]}{single_sheet_i}'] = lesson
                        if max_width < len(lesson):
                            max_width = len(lesson)
                    except:
                        worksheets_dict[char_index + 1][f'{characters[single_sheet_char_index]}{single_sheet_i}'] = ''

                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].font = FONT_LESSONS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].border = BORDER_LESSONS
                    worksheets_dict[char_index + 1][
                        f'{characters[single_sheet_char_index]}{single_sheet_i}'].alignment = ALIGNMENT

                    if hours[hour_index] != '16.00-17.00':
                        hour_index += 1
                    else:
                        insert_day = True
                        hour_index = 0
                        day_index += 1
                        if day_index == 5:
                            break
                single_sheet_i += 1
            worksheets_dict[char_index + 1].column_dimensions[characters[single_sheet_char_index]].width = max_width - 5
            ws1 = wb.active
        char_index += 1
    ws1.row_dimensions[1].height = 25
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    response = HttpResponse(content=stream, content_type='application/ms-excel', )
    response['Content-Disposition'] = 'attachment; filename=orario_settimanale.xlsx'
    return response


def copy_old_schedule(request):
    ref_date = datetime.now().date()
    course_list = Course.objects.filter(is_finished=False)
    while True:
        if ref_date.strftime("%A") == 'lunedì':
            break
        else:
            ref_date = ref_date - timedelta(days=1)
    date_range = [ref_date, ref_date + timedelta(days=4)]
    old_lessons = LessonWithDate.objects.filter(lesson_date__range=date_range).order_by('lesson_date')
    while True:
        if old_lessons:
            break
        else:
            date_range[0] = date_range[0] - timedelta(weeks=1)
            date_range[1] = date_range[1] - timedelta(weeks=1)
            old_lessons = LessonWithDate.objects.filter(lesson_date__range=date_range)
    lessons_to_reset = LessonWithDate.objects.filter(lesson_date__gte=old_lessons.first().lesson_date+timedelta(weeks=1))
    for lesson in lessons_to_reset:
        lesson.delete()
    for lesson in old_lessons:
        new_lesson = LessonWithDate(lesson_date=lesson.lesson_date+timedelta(weeks=1),
                                    hour=lesson.hour,
                                    training_unit=lesson.training_unit,
                                    classroom=lesson.classroom)
        try:
            new_lesson.save()
        except:
            print('Lezione già presente')
    return render(request, 'lesson_with_date/lesson_with_date_list.html', {'course_list': course_list,
                                                                           'ref_date': ref_date,
                                                                           'next_week': (ref_date + timedelta(
                                                                               weeks=1)),
                                                                           'message': 'Hai copiato l`ultimo orario in quello nella prossima per tutti i corsi'
                                                                           })

class ValidateExtraLessonView(LoginRequiredMixin, CreateView):
    permission_required = 'frequenze.add_lessondone'
    model = LessonDone
    form_class = LessonDoneModelForm
    template_name = "lesson_done/lesson_done_create.html"
    success_url = "/"


def lesson_done_list_view(request):
    permission = request.user.has_perm('frequenze.view_lessondone')
    if permission:
        lessons = LessonDone.objects.all().order_by('-lesson_date', 'training_unit__course__name', 'hour')
        context = {}
        training_units = TrainingUnit.objects.all().order_by('course__name', 'code')
        teachers = Staff.objects.all().order_by('surname')
        courses = Course.objects.all().order_by('name')
        if request.method == 'POST':
            parameters = {'start_date': request.POST.get('startDate'),
                          'end_date': request.POST.get('endDate'),
                          'course': request.POST.get('course'),
                          'teacher': request.POST.get('teacher'),
                          'training_unit': request.POST.get('uf'),
                          }

            # parameters = {k: v for k, v in parameters.items() if v} # rimuove i parametri vuoti
            if parameters['start_date'] and parameters['end_date']:
                lessons = lessons.filter(lesson_date__range=[parameters['start_date'], parameters['end_date']])
            elif parameters['start_date']:
                lessons = lessons.filter(lesson_date__gte=parameters['start_date'])
            elif parameters['end_date']:
                lessons = lessons.filter(lesson_date__lte=parameters['end_date'])
            if parameters['training_unit']:
                lessons = lessons.filter(training_unit=parameters['training_unit'])
            else:
                if parameters['course']:
                    lessons = lessons.filter(training_unit__course=parameters['course'])
                if parameters['teacher']:
                    lessons = lessons.filter(training_unit__contract_related_unit__employee=parameters['teacher'])

        context = {'lessons': lessons,
                   'training_units': training_units,
                   'teachers': teachers,
                   'courses': courses
                   }
        return render(request, 'lesson_done/lesson_done_list.html', context)
    else:
        return render(request, 'missing_auth.html')
