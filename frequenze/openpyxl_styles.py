from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment

FONT_DAYS = Font(name='Calibri',
                 size=10,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000',
                 )
FONT_COURSES = Font(name='Calibri',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='004a53',
                    )
FONT_HOURS = Font(name='Calibri',
                  size=8,
                  bold=False,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='191400',
                  )
FONT_LESSONS = Font(name='Calibri',
                    size=8,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000',
                    )

FILL_DAYS = PatternFill("solid", fgColor="ffe680")

FILL_COURSES = PatternFill("solid", fgColor="b2eaf1")

FILL_L = PatternFill("solid", fgColor="ffd1d1")
FILL_U = PatternFill("solid", fgColor="ebffff")
FILL_T = PatternFill("solid", fgColor="ffffd9")



BORDER_DAYS = Border(left=Side(border_style=None,
                               color='FF000000'),
                     right=Side(border_style=None,
                                color='FF000000'),
                     top=Side(border_style='double',
                              color='FF000000'),
                     bottom=Side(border_style='double',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000')
                     )

BORDER_COURSES = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style=None,
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000')
                     )


BORDER_HOURS = Border(left=Side(border_style='thin',
                                color='FF000000'),
                      right=Side(border_style='thin',
                                 color='FF000000'),
                      top=Side(border_style='thin',
                               color='FF000000'),
                      bottom=Side(border_style='thin',
                                  color='FF000000'),
                      diagonal=Side(border_style=None,
                                    color='FF000000'),
                      diagonal_direction=0,
                      outline=Side(border_style=None,
                                   color='FF000000'),
                      vertical=Side(border_style=None,
                                    color='FF000000'),
                      horizontal=Side(border_style=None,
                                      color='FF000000')
                      )

BORDER_LESSONS = Border(left=Side(border_style='medium',
                                  color='FF000000'),
                        right=Side(border_style='medium',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style=None,
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style=None,
                                     color='FF000000'),
                        vertical=Side(border_style=None,
                                      color='FF000000'),
                        horizontal=Side(border_style=None,
                                        color='FF000000')
                        )

ALIGNMENT = Alignment(horizontal='center',
                      vertical='center',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)
